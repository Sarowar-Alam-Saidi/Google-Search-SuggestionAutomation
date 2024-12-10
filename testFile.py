import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


# Function to find the longest and shortest options from suggestions
def find_longest_and_shortest_options(suggestions):
    longest_option = max(suggestions, key=len)
    shortest_option = min(suggestions, key=len)
    return longest_option, shortest_option


# Function to read keywords from today's sheet in the Excel file
def get_keywords_from_excel(file_path):
    today = datetime.datetime.now().strftime('%A')  # Get today's weekday (e.g., Monday)
    workbook = openpyxl.load_workbook(file_path)

    if today not in workbook.sheetnames:
        raise ValueError(f"No sheet named '{today}' found in the Excel file.")

    sheet = workbook[today]
    keywords = [sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1) if
                sheet.cell(row=row, column=1).value]
    return keywords, workbook, sheet


# Function to update Excel with longest and shortest options
def update_excel(sheet, row, longest, shortest):
    sheet.cell(row=row, column=2).value = longest
    sheet.cell(row=row, column=3).value = shortest


# Main function to process keywords
def process_keywords(file_path):
    # Initialize WebDriver with automatic management
    driver = webdriver.Chrome()  # Selenium will handle the driver automatically
    driver.get("https://www.google.com")

    # Accept cookies if the consent banner appears (adjust selector as per region)
    try:
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "L2AGLb"))).click()
    except Exception:
        pass

    # Load keywords from Excel
    keywords, workbook, sheet = get_keywords_from_excel(file_path)

    for index, keyword in enumerate(keywords, start=2):  # Start from row 2
        # Search the keyword on Google
        search_box = driver.find_element(By.NAME, "q")
        search_box.clear()
        search_box.send_keys(keyword)

        # Wait explicitly for the suggestions to load
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".G43f7e"))
            )
            suggestions_elements = driver.find_elements(By.CSS_SELECTOR, ".G43f7e")
            suggestions = [element.text for element in suggestions_elements if element.text.strip()]

            if suggestions:
                # Determine longest and shortest suggestions
                longest, shortest = find_longest_and_shortest_options(suggestions)

                # Update Excel sheet
                update_excel(sheet, index, longest, shortest)
            else:
                print(f"No suggestions found for keyword: {keyword}")
        except Exception as e:
            print(f"An error occurred while fetching suggestions for keyword '{keyword}': {e}")

        # Clear the search box for the next keyword
        search_box.clear()

    # Save the updated Excel file
    workbook.save(file_path)

    # Close the browser
    driver.quit()


# Specify the path to your Excel file
file_path = "testxlfile.xlsx"  # Replace with the actual path to your file

# Execute the script
try:
    process_keywords(file_path)
    print("Process completed successfully. The Excel file has been updated.")
except Exception as e:
    print(f"An error occurred: {e}")
